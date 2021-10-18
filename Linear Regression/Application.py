from tkinter import*
from openpyxl import*
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from tkinter import messagebox
import pandas as pd
data=pd.read_excel("C:\\Users\\HP Elitebook G6\\Desktop\\Python\\MLalgorythm\\Linear Regression\\database.xlsx")
#print(data)

def cleardata():
    ans.configure(text="0")
        
def pushdata(user_age,result):
    #Package for inserting data into an excel server
    
    
    #Step 1:- Open the excel server
    wb=load_workbook("C:\\Users\\HP Elitebook G6\\Desktop\\Python\\MLalgorythm\\Linear Regression\\database.xlsx")
    
    #Step 2:- Enable the excel 
    sheet=wb.active
    
    #Steap 3:- Insert the data
    age=user_age
    weight=result
    
    current_row=sheet.max_row
    print(current_row)
    
    sheet.cell(row=current_row+1, column=1).value=age
    sheet.cell(row=current_row+1, column=2).value=weight
    
    #Step 4:- Save the excel file
    wb.save("C:\\Users\\HP Elitebook G6\\Desktop\\Python\\MLalgorythm\\Linear Regression\\database.xlsx")
    print("done")
    messagebox.showinfo('info','database updated sucessfully')
    
    
def predict():
    #Fetch the age from the interface
    user_age=int(s1.get())
    #Apply the logic
    # Apply the linear regression on age and weight

    #import the mathamatical module from 
    from scipy import stats #it is called stats module 
    import matplotlib.pyplot as plt
    def calculate(x):
        return slope*x+intercept


    x=(age_list)
    y=(weight_list)
    
    #calculate the slope and intercept
    slope,intercept,r,p,sd=stats.linregress(x,y)
    #apply the algorythm
    model=list(map(calculate,x))
    plt.plot(x,y)
    
    #calculate the result for given value of x
    nx=[0,user_age]
    #Predict the value
    ny=list(map(calculate,nx))
    result=ny[1]
    print(result)
    #Push the result into backend database
    pushdata(user_age,result)
    ans.configure(text=result,font=800,fg='green')
    figure=plt.Figure(figsize=(6,5),dpi=100)
    ax=figure.add_subplot()
    bargraph=FigureCanvasTkAgg(figure,application)
    Graph=bargraph.get_tk_widget()
    Graph.place(x=700, y=160)
    df=pd.DataFrame({'Age':age_list,'Weight':weight_list})
    df.plot(x='Age',y='Weight', color='red', legend=True,ax=ax)
    plt.plot(nx,ny,color="red")
    plt.show()



#Exctract the age. 
age=data['Age']
age_list=list(age)
#print(age_list)

weight=data['Weight']
weight_list=list(weight)
#print(weight_list)

#Creat an application
application=Tk()
application.geometry('1550x1550')

#Creat a lable
l1=Label(application,text="BRM Calculator", font=500)
l1.place(x=700,y=60)

#Creat a gender button
g1=Label(application,text="Choose your gender", font=300)
g1.place(x=300,y=120)

#Creat a radio button
r1=Radiobutton(application,text="Male",font=300)
r1.place(x=320,y=150)

r2=Radiobutton(application,text="Female",font=300)
r2.place(x=320,y=180)

#Creat a lable
l2=Label(application,text="Choose age", font=300)
l2.place(x=300,y=240)


#Creat a spinbox
s1=Spinbox(application,from_=1,to=100)
s1.place(x=320,y=280)


#Creat a predict putton
b1=Button(application,text="Predict",command=predict)
b1.place(x=340,y=320)


ans=Label(application,text="0", font=1000)
ans.place(x=380,y=380)

#Creat a clear button
C1=Button(application,text="Clear data",command=cleardata)
C1.place(x=400,y=320)

application.mainloop()