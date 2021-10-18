#Package for inserting data into an excel server
from openpyxl import*

#Step 1:- Open the excel server
wb=load_workbook("C:\\Users\\HP Elitebook G6\\Desktop\\Python\\MLalgorythm\\Linear Regression\\database.xlsx")

#Step 2:- Enable the excel 
sheet=wb.active

#Steap 3:- Insert the data
age=55
weight=95

current_row=sheet.max_row
print(current_row)

sheet.cell(row=current_row+1, column=1).value=age
sheet.cell(row=current_row+1, column=2).value=weight

#Step 4:- Save the excel file
wb.save("C:\\Users\\HP Elitebook G6\\Desktop\\Python\\MLalgorythm\\Linear Regression\\database.xlsx")
print("done")